import sys
import openpyxl
from openpyxl import Workbook
import openpyxl.styles.colors
from copy import copy
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


def load_substitutions(filename):
    dct = {}
    with open(filename) as fp:
        line = fp.readline()
        while line:
            pair = line.split(',')
            dct[pair[0].strip()] = pair[1].strip()
            line = fp.readline()
    return dct


def is_back_label(lbl):
    return len(lbl) > 0 and lbl[0] == '<'


def is_link_label(lbl):
    return len(lbl) > 0 and lbl[0] == '*'


def is_textlink_label(lbl):
    return len(lbl) > 0 and lbl[0] == '|'


def is_kbd_label(lbl):
    return len(lbl) > 0 and lbl[0] == '[' and lbl[len(lbl)-1] == ']'


def is_pred_label(lbl):
    return lbl == "{pred}"


def is_inflector_label(lbl):
    return lbl[0] == '{' and lbl[-1] == '}' and not is_pred_label(lbl)


def infl_from_inflector_label(lbl):
    txt = lbl[1:-1]
    if "|" not in txt:
        return txt
    parts = txt.split("|")
    return parts[0]


def link_from_inflector_label(lbl):
    txt = lbl[1:-1]
    if "|" not in txt:
        return None
    parts = txt.split("|")
    return parts[1]


def link_from_label(lbl):
    return lbl[1:len(lbl)]


def link_from_textlink_label(lbl):
    txtnlnk = lbl.split("|")
    return txtnlnk[2]


def text_from_textlink_label(lbl):
    txtnlnk = lbl.split("|")
    return txtnlnk[1]


def dont_morph(lbl):
    return len(lbl) <= 1 or lbl[-1] != '#'


def demorph_label(lbl):
    if dont_morph(lbl):
        return lbl
    return lbl[:-1]


def text_from_kbd(lbl):
    return lbl[1:len(lbl)-1]


def capitalize(lbl):
    if not lbl or len(lbl) < 1:
        return lbl
    return lbl[0].upper() + lbl[1:]


def transform_label(lbl, xforms):
    domorph = False
    if not dont_morph(lbl):
        domorph = True
        lbl = demorph_label(lbl)

    if lbl in xforms:
        lbl = xforms[lbl]
    elif lbl[0].isupper():
        llbl = lbl.lower()
        if llbl in xforms:
            lbl = capitalize(xforms[llbl])

    if domorph:
        lbl = lbl + "#"
    return lbl


def transform_value(lbl, xforms):
    if lbl == "color":
        x = 1
    if not lbl or len(lbl) == 0 or is_pred_label(lbl) or is_kbd_label(lbl) or is_back_label(lbl):
        return lbl

    if is_link_label(lbl):
        lnk = link_from_label(lbl)
        return "*" + transform_label(lnk, xforms)

    if is_inflector_label(lbl):
        lnk = link_from_inflector_label(lbl)
        if not lnk:
            return lbl
        lnk = transform_label(lnk, xforms)
        return "{" + infl_from_inflector_label(lbl) + "|" + lnk + "}"

    if is_textlink_label(lbl):
        lnk = link_from_textlink_label(lbl)
        txt = text_from_textlink_label(lbl)
        return "|" + transform_label(txt, xforms) + "|" + transform_label(lnk, xforms)

    return transform_label(lbl, xforms)


subs = load_substitutions(sys.argv[3])
wbi = openpyxl.load_workbook(sys.argv[1])
wbo = Workbook()
for wsi in wbi:
    wso = wbo.create_sheet(title=wsi.title)
    print(f"{wsi.title}\n")
    for xcol in range(1, wsi.max_column + 1):
        for xrow in range(1, wsi.max_row + 1):
            ci = wsi.cell(row=xrow, column=xcol)
            value = transform_value(ci.value, subs)
            co = wso.cell(row=xrow, column=xcol, value=value)
            co.font = copy(ci.font)
            co.fill = copy(ci.fill)
            co.alignment = copy(ci.alignment)
            co.border = copy(ci.border)

wbo.save(sys.argv[2])

